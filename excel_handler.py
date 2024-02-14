import openpyxl


class ExcelHandler:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path)

    def read_card_names(self, sheet_name: str = 'Template', column: str = 'A', start_row: int = 3) -> list[str]:
        """
        Read card names from an Excel file.

        Args:
            sheet_name (str, optional): The name of the sheet to read from. Defaults to 'Template'.
            column (str, optional): The column to read from. Defaults to 'A'.
            start_row (int, optional): The starting row to read from. Defaults to 3.

        Returns:
            list: A list of card names.
        """
        if sheet_name not in self.workbook.sheetnames:
            sheet = self.workbook.active
        else:
            sheet = self.workbook[sheet_name]
        first_column_values = [cell.value for cell in sheet[column][start_row:] if cell.value is not None]
        return first_column_values

    def fill_in_column(self, list_of_values: list[float], sheet_name: str, column, start_row: int = 3) -> None:
        """
        Fills in a column in an Excel file with a list of values.

        Args:
            list_of_values (list[float]): The list of values to be written to the column.
            sheet_name (str): The name of the sheet in the Excel file.
            column (str, optional): The column letter to write the values to.
            start_row (int, optional): The starting row number to write the values to. Defaults to 3.
        """
        if sheet_name not in self.workbook.sheetnames:
            print('Sheet is not found. Not writing to excel.')
            return
        else:
            sheet = self.workbook[sheet_name]

        for i, value in enumerate(list_of_values, start=1):
            sheet.cell(row=start_row + i, column=column, value=value)

    def copy_sheet(self, source_sheet_name='Template', destination_sheet_name='Y') -> None:
        """
        Copies a sheet from an Excel file to another sheet with a specified name.

        Args:
            source_sheet_name (str, optional): The name of the source sheet to be copied. Defaults to 'Template'.
            destination_sheet_name (str, optional): The name of the destination sheet. Defaults to 'Y'.
        """
        if source_sheet_name not in self.workbook.sheetnames:
            print('Source sheet not found. Cannot copy.')
            return
        if destination_sheet_name in self.workbook.sheetnames:
            print('Destination sheet already exists. Cannot copy.')
            return

        source_sheet = self.workbook[source_sheet_name]
        destination_sheet = self.workbook.copy_worksheet(source_sheet)
        destination_sheet.title = destination_sheet_name

        # Reorder sheets
        sheets = self.workbook.sheetnames
        sheets.remove(destination_sheet_name)
        sheets.insert(2, destination_sheet_name)
        self.workbook._sheets = [self.workbook[s] for s in sheets]

    def update_changes_sheet(self, new_sheet_name, sheet_name: str = 'Changes', card_name_list: list[str] = None,
                             from_to_row: int = 1, from_column: int = 3, to_column: int = 4,
                             card_name_column: int = 1, card_name_start_row: int = 3) -> None:
        """
        'Changes' is the first sheet that compares prices two period sheets. This function updates pointers to comparison sheets in C1/D1 cell,
        and optionally updates the card name list if new cards were added in the 'Template' sheet.

        Args:
            new_sheet_name (str): The name of the new sheet with latest prices (obtained from 'sheet_name_now' var in main.py)
            sheet_name (str, optional): The name of the changes sheet. Defaults to 'Changes'.
            card_name_list (list[str], optional): The list of card names. Defaults to None. Optional.
            from_to_row (int, optional): The row number to write the new sheet name to. Defaults to 1.
            from_column (int, optional): Where name of the older sheet is. Defaults to 3. (C1 cell)
            to_column (int, optional): Where name of the newer sheet is. Compares values from older with newer. Defaults to 4. (D1 cell)
            card_name_column (int, optional): The column to write the card names to. Defaults to 1 (A).
            card_name_start_row (int, optional): The starting row to write the card names to. Defaults to 3.
        """
        if sheet_name not in self.workbook.sheetnames:
            print(f'"{sheet_name}" sheet is not found. Not writing to excel.')
            return
        else:
            sheet = self.workbook[sheet_name]

        # NOTE: "compare prices from Feb 10 to Feb 14" -> Feb 10 is "old_from_period_sheet" name, Feb 14 is "old_to_sheet_name"
        old_to_sheet_name = sheet.cell(row=from_to_row, column=to_column).value

        sheet.cell(row=from_to_row, column=from_column, value=old_to_sheet_name)  # setting C1 cell "Compare from" to old "Compare to"
        sheet.cell(row=from_to_row, column=to_column, value=new_sheet_name)  # setting D1 cell "Compare to" to new sheet created in main.py
        # NOTE: Before: "Compare Feb 10 to Feb 14" -> After: "Compare Feb 14 to 'new_sheet_name'"

        # Update card name list
        if card_name_list:
            for i, value in enumerate(card_name_list, start=1):
                sheet.cell(row=card_name_start_row + i, column=card_name_column, value=value)

    def save_and_close(self):
        self.workbook.save(self.excel_path)
        self.workbook.close()
